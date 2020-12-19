from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask, request
import json
import datetime

file      = 'data_julia.xlsx'
cash      = []
cash_size = 5

app = Flask(__name__)
@app.route('/', methods=['POST'])
def index():
    global cash_size, file, cash

    data = request.json

    for msg in data['check']:
        cash.append( [data['user_id'],
                        str(datetime.datetime.now()),
                        msg['item'],
                        msg['price']
                        ])
        if len(cash) >= cash_size:
            try:
                w_book  = load_workbook(file)
                w_sheet = w_book.active
                w_book.save(file)
                w_book.close()

            except FileNotFoundError:
                w_book = Workbook()
                w_sheet = w_book.active
                w_sheet.append(['N', 'User ID', 'Date time', 'Item', 'Price'])
                w_book.save(file)

            except Exception as e:
                print('Произошла ошибка!', e)
                return "Ok"

            w_book = load_workbook(file)
            w_sheet = w_book.active

            last_raw = w_sheet.max_row

            for new_raw in cash:
                new_raw=[last_raw]+new_raw
                w_sheet.append(new_raw)
                last_raw+=1
            w_book.save(file)
            w_book.close()
            cash=[]

    return "Ok"

if __name__ == "__main__":
    app.run()
